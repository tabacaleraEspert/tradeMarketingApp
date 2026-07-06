"""
Export de todos los PDV con su TMR asignado, sus rutas y todas sus notas.

Genera un .xlsx (y un .csv) con una fila por nota. Los PDV sin notas
aparecen igual con las columnas de nota vacías (LEFT JOIN).

OJO: el TMR del PDV es `PDV.AssignedUserId` (se hereda al incluir el PDV
en una ruta). NO confundir con "Nota creada por", que es quién escribió
cada comentario (vacío si el PDV no tiene notas).

El .xlsx trae 2 hojas:
  - "PDV + TMR + Notas": todo el detalle.
  - "PDV sin TMR": solo los PDV con AssignedUserId NULL (para revisar).

Uso:
    python scripts/export_pdv_notes.py
"""
import csv
import os
from datetime import datetime, timezone

import pymssql
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SERVER = os.environ.get("DATABASE_SERVER", "trade-mkt-sql.database.windows.net")
DB = os.environ.get("DATABASE_NAME", "trademktdb")
USER = os.environ.get("DATABASE_USER", "tmadmin")
PWD = os.environ.get("DATABASE_PASSWORD", "TradeMkt2026Azr99")

QUERY = """
SELECT
    p.PdvId,
    p.Code              AS PdvCode,
    p.Name              AS PdvName,
    p.BusinessName      AS PdvBusinessName,
    p.Address           AS PdvAddress,
    p.City              AS PdvCity,
    z.Name              AS Zone,
    p.IsActive          AS PdvActiva,
    tmr.DisplayName     AS TmrAsignado,
    tmr.Email           AS TmrEmail,
    rutas.Rutas         AS Rutas,
    n.PdvNoteId,
    n.Content           AS NotaContenido,
    n.CreatedAt         AS NotaCreadaEl,
    creator.DisplayName AS NotaCreadaPor,
    creator.Email       AS NotaCreadaPorEmail,
    n.IsResolved        AS NotaResuelta,
    resolver.DisplayName AS NotaResueltaPor,
    n.ResolvedAt        AS NotaResueltaEl,
    n.VisitId           AS VisitaId
FROM PDV p
LEFT JOIN [User] tmr     ON tmr.UserId = p.AssignedUserId
LEFT JOIN (
    SELECT rp.PdvId, STRING_AGG(r.Name, ' | ') AS Rutas
    FROM RoutePdv rp JOIN Route r ON r.RouteId = rp.RouteId
    GROUP BY rp.PdvId
) rutas ON rutas.PdvId = p.PdvId
LEFT JOIN PdvNote n      ON n.PdvId = p.PdvId
LEFT JOIN [User] creator ON creator.UserId = n.CreatedByUserId
LEFT JOIN [User] resolver ON resolver.UserId = n.ResolvedByUserId
LEFT JOIN Zone z         ON z.ZoneId = p.ZoneId
ORDER BY p.Name, n.CreatedAt;
"""

# Solo los PDV que NO tienen TMR asignado (para hoja de revisión)
QUERY_SIN_TMR = """
SELECT
    p.PdvId, p.Code, p.Name, p.BusinessName, p.City, z.Name AS Zone,
    p.IsActive,
    CASE WHEN rp.PdvId IS NULL THEN 'No' ELSE 'Si' END AS EnRuta,
    p.CreatedAt
FROM PDV p
LEFT JOIN Zone z ON z.ZoneId = p.ZoneId
LEFT JOIN (SELECT DISTINCT PdvId FROM RoutePdv) rp ON rp.PdvId = p.PdvId
WHERE p.AssignedUserId IS NULL
ORDER BY p.IsActive DESC, p.Name;
"""

HEADERS = [
    "PdvId", "Codigo PDV", "Nombre PDV", "Razon Social", "Direccion", "Ciudad",
    "Zona", "PDV Activa",
    "TMR asignado", "Email TMR", "Rutas",
    "NotaId", "Nota (contenido)", "Nota creada el", "Nota creada por",
    "Email creador", "Nota resuelta", "Resuelta por", "Resuelta el", "VisitaId",
]

HEADERS_SIN_TMR = [
    "PdvId", "Codigo PDV", "Nombre PDV", "Razon Social", "Ciudad", "Zona",
    "PDV Activa", "En ruta", "Creado el",
]


def fmt(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, bool):
        return "Si" if v else "No"
    return v


def main():
    print(f"Conectando a {SERVER}/{DB} ...")
    conn = pymssql.connect(server=SERVER, user=USER, password=PWD, database=DB, login_timeout=90)
    cur = conn.cursor()
    cur.execute(QUERY)
    rows = cur.fetchall()
    cur.execute(QUERY_SIN_TMR)
    rows_sin_tmr = cur.fetchall()
    cur.close()
    conn.close()

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    out_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(out_dir, f"PDV_Notas_Espert_{stamp}.xlsx")
    csv_path = os.path.join(out_dir, f"PDV_Notas_Espert_{stamp}.csv")

    # ---- XLSX ----
    wb = Workbook()
    ws = wb.active
    ws.title = "PDV + Notas"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([fmt(v) for v in r])

    # Ancho de columnas
    widths = [8, 14, 34, 28, 30, 16, 16, 9, 22, 30, 30, 8, 60, 16, 22, 28, 12, 22, 16, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Hoja 2: PDV sin TMR ----
    ws2 = wb.create_sheet("PDV sin TMR")
    ws2.append(HEADERS_SIN_TMR)
    for c in range(1, len(HEADERS_SIN_TMR) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    ws2.freeze_panes = "A2"
    for r in rows_sin_tmr:
        ws2.append([fmt(v) for v in r])
    for i, w in enumerate([8, 20, 30, 26, 16, 18, 9, 8, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(xlsx_path)

    # ---- CSV (hoja principal) ----
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in rows:
            w.writerow([fmt(v) for v in r])

    NOTE_IDX = HEADERS.index("NotaId")        # col de NotaId (para contar notas)
    TMR_IDX = HEADERS.index("TMR asignado")   # col de TMR asignado
    notes = sum(1 for r in rows if r[NOTE_IDX] is not None)
    pdvs = len({r[0] for r in rows})
    pdv_con_tmr = len({r[0] for r in rows if r[TMR_IDX] is not None})
    print(f"Filas: {len(rows)}  |  PDVs: {pdvs}  |  Notas: {notes}")
    print(f"PDV con TMR: {pdv_con_tmr}  |  PDV sin TMR: {pdvs - pdv_con_tmr}")
    print(f"XLSX: {xlsx_path}  (hoja 2 = 'PDV sin TMR': {len(rows_sin_tmr)} filas)")
    print(f"CSV:  {csv_path}")


if __name__ == "__main__":
    main()
